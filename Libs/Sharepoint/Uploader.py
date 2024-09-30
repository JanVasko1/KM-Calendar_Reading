#into je tu: D:\Time_Sheet_App
from Libs.Sharepoint.Authentication import Init_authentication
from openpyxl import load_workbook
from pandas import DataFrame
import pandas
import sharepy
import json
from tqdm import tqdm

# ---------------------------------------------------------- Set Defaults ---------------------------------------------------------- #
File = open(file=f"Libs\\Settings.json", mode="r", encoding="UTF-8", errors="ignore")
Settings = json.load(fp=File)
File.close()

SP_Link = Settings["General"]["Downloader"]["Sharepoint"]["Link"]
User_Email = Settings["General"]["Downloader"]["Sharepoint"]["Auth"]["Email"]


def data_Get(ws, data_boundary):
    data = ws[data_boundary]
    content = [[str(cell.internal_value) for cell in ent] for ent in data]
    header = content[0]
    data_rows = content[1:]
    TimeSheets_df = pandas.DataFrame(data=data_rows, columns = header)
    TimeSheets_df = TimeSheets_df[["Personnel number", "Date", "Network Description", "Activity", "Activity description", "Start Time", "End Time", "Location"]]
    return TimeSheets_df

def Timesheets_Identify_empty_row(TimeSheets_df: DataFrame) -> list[str, str]:
    # Reorder lines
    TimeSheets_df2 = TimeSheets_df.sort_index(ascending=False)

    # Find first non empty line and get index of empty line (crossing from NULL to Value)
    Pre_Row_Personnel_number = "Row_Personnel_number"
    Pre_Row_Date = "Row_Date"
    Pre_Row_Network_Description = "Row_Network_Description"
    Pre_Row_Activity = "Row_Activity"
    Pre_Row_Activity_description = "Row_Activity_description"
    Pre_Row_Start_Time = "Row_Start_Time"
    Pre_Row_End_Time = "Row_End_Time"

    for row in TimeSheets_df2.iterrows():
        # read values from row
        Row_Series = pandas.Series(row[1])
        Row_Index = row[0]
        Row_Personnel_number = Row_Series["Personnel number"]
        Row_Date = Row_Series["Date"]
        Row_Network_Description = Row_Series["Network Description"]
        Row_Activity = Row_Series["Activity"]
        Row_Activity_description = Row_Series["Activity description"]
        Row_Start_Time = Row_Series["Start Time"]
        Row_End_Time = Row_Series["End Time"]

        # Compare
        if (Pre_Row_Personnel_number == "None") and (Pre_Row_Date == "None") and (Pre_Row_Network_Description == "None") and (Pre_Row_Activity == "None") and (Pre_Row_Activity_description == "None") and (Pre_Row_Start_Time == "None") and (Pre_Row_End_Time == "None"):
            if (Row_Personnel_number == "None") and (Row_Date == "None") and (Row_Network_Description == "None") and (Row_Activity == "None") and (Row_Activity_description == "None") and (Row_Start_Time == "None") and (Row_End_Time == "None"):
                Excel_row_No = Row_Index
            else:
                Excel_row_No += 2   # add 2 because of that program works with indexes (start from 0) and find first occurence of values
                break

        Pre_Row_Personnel_number = Row_Personnel_number
        Pre_Row_Date = Row_Date
        Pre_Row_Network_Description = Row_Network_Description
        Pre_Row_Activity = Row_Activity
        Pre_Row_Activity_description = Row_Activity_description
        Pre_Row_Start_Time = Row_Start_Time
        Pre_Row_End_Time = Row_End_Time

    # Check because of first line
    if Excel_row_No == 0:
        Excel_row_No = 2
    else:
        pass
    
    A_Cell = f"A{Excel_row_No}"
    E_Cell = f"E{Excel_row_No}"
    return A_Cell, E_Cell

# ---------------------------------------------------------- Main Function ---------------------------------------------------------- #
def Upload(Events: DataFrame) -> None:
    # Connect to Sharepoint and Authnetify
    while True:
        # Authorisation
        try:
            s_aut = sharepy.load(filename=f".\\Libs\\Sharepoint\\ssp-session.pkl")
        except:
            s_aut = Init_authentication()

        # Check loop + check auth return value
        if s_aut == "":
            print("Not authenticated. Wrong password or try connect to Baracuda. ")
            Break = input(f"Do you want to stop loop? [Y/N]?")
            Break = Break.upper()
            if Break == "Y":
                break
            else:
                pass
        else:
            break

    # Download
    r = s_aut.getfile(f"{SP_Link}", filename=".\\Libs\\Sharepoint\\TimeSheets.xlsx")

    # Read downloaded data
    wb = load_workbook(filename=".\\Libs\\Sharepoint\\TimeSheets.xlsx")
    ws = wb["TimeSpent"]

    Table_list = []
    for key, value in ws.tables.items():
        Table = []
        Table.append(str(key))
        Table.append(str(value))
        Table_list.append(Table)

    # TimeSheets_df
    data_boundary = Table_list[0][1]
    data_boundary = data_boundary.replace("O", "J")
    TimeSheets_df = data_Get(ws=ws, data_boundary=data_boundary)
    A_Cell, E_Cell = Timesheets_Identify_empty_row(TimeSheets_df=TimeSheets_df)
    print(f"First Cell: {A_Cell}, {E_Cell}")
    #! Dodělat !!!
    #! Stáhnout si i obsah toho co je už v excelu vloženo (mít ot v dataframe) a porovnat s tím co už mám vložený s tím co chci vkládat a za sebe vložit pouze rozdíl --> to způsobí to, že můžu nechat