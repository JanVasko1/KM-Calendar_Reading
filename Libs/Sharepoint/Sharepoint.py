# Import Libraries
from openpyxl import load_workbook
from pandas import DataFrame, Series, read_excel
import sharepy

import Libs.GUI.Elements as Elements
import Libs.Sharepoint.Authentication as Authentication
import Libs.Data_Functions as Data_Functions

# ---------------------------------------------------------- Local Functions ---------------------------------------------------------- #
def Get_Table_Data(ws, data_boundary) -> DataFrame:
    data = ws[data_boundary]
    content = [[str(cell.internal_value) for cell in ent] for ent in data]
    header = content[0]
    data_rows = content[1:]
    TimeSheets_df = DataFrame(data=data_rows, columns = header)
    TimeSheets_df = TimeSheets_df[["Personnel number", "Date", "Network Description", "Activity", "Activity description", "Start Time", "End Time", "Location"]]
    return TimeSheets_df

def Download_Excel(Settings: dict, s_aut: sharepy, SP_Link: str, Type: str, Name: str|None) -> str:
    SP_Link_domain = Settings["0"]["General"]["Downloader"]["Sharepoint"]["Auth"]["Auth_Address"]
    SP_File_Name = Settings["0"]["General"]["Downloader"]["Sharepoint"]["File_name"]

    # Download
    if Type == "Current":
        response = s_aut.getfile(f"{SP_Link_domain}{SP_Link}", filename=Data_Functions.Absolute_path(relative_path=f"Operational\\Downloads\\{SP_File_Name}.xlsm"))
    elif Type == "History":
        response = s_aut.getfile(f"{SP_Link_domain}{SP_Link}", filename=Data_Functions.Absolute_path(relative_path=f"Operational\\History\\{Name}.xlsm"))
    elif Type == "Team":
        response = s_aut.getfile(f"{SP_Link_domain}{SP_Link}", filename=Data_Functions.Absolute_path(relative_path=f"Operational\\My_Team\\{Name}.xlsm"))
    else:
        return False

    if response.status_code == 200:
        return True
    else:
        return False

def Get_WorkSheet(Settings: dict, Sheet_Name: str, Type: str, Name: str|None):
    SP_File_Name = Settings["0"]["General"]["Downloader"]["Sharepoint"]["File_name"]

    if Type == "Current":
        WorkBook = load_workbook(filename=Data_Functions.Absolute_path(relative_path=f"Operational\\Downloads\\{SP_File_Name}.xlsm"))
    elif Type == "History":
        WorkBook = load_workbook(filename=Data_Functions.Absolute_path(relative_path=f"Operational\\History\\{Name}.xlsm"))
    elif Type == "Team":
        WorkBook = load_workbook(filename=Data_Functions.Absolute_path(relative_path=f"Operational\\My_Team\\{Name}.xlsm"))
    Sheet = WorkBook[Sheet_Name]
    return Sheet

def Get_Tables_on_Worksheet(Sheet) -> list:
    Table_list = []
    for key, value in Sheet.tables.items():
        Table = []
        Table.append(str(key))
        Table.append(str(value))
        Table_list.append(Table)
    return Table_list
    

def Time_sheets_Identify_empty_row(TimeSheets_df: DataFrame) -> list[str, str]:
    # Reorder lines
    TimeSheets_df2 = TimeSheets_df.sort_index(ascending=False)

    # Find first non empty line and get index of empty line (crossing from NULL to Value)
    Pre_Row_Personnel_number = "Row_Personnel_number"
    Pre_Row_Date = "Row_Date"
    Pre_Row_Network_Description = "Row_Network_Description"
    Pre_Row_Activity = "Row_Activity"
    Pre_Row_Activity_description = "Row_Activity_description"
    Pre_Row_Start_Time = "Row_Start_Time"
    Pre_Row_End_Time = "Row_End_Time"

    for row in TimeSheets_df2.iterrows():
        # read values from row
        Row_Series = Series(row[1])
        Row_Index = row[0]
        Row_Personnel_number = Row_Series["Personnel number"]
        Row_Date = Row_Series["Date"]
        Row_Network_Description = Row_Series["Network Description"]
        Row_Activity = Row_Series["Activity"]
        Row_Activity_description = Row_Series["Activity description"]
        Row_Start_Time = Row_Series["Start Time"]
        Row_End_Time = Row_Series["End Time"]

        # Compare
        if (Pre_Row_Personnel_number == "None") and (Pre_Row_Date == "None") and (Pre_Row_Network_Description == "None") and (Pre_Row_Activity == "None") and (Pre_Row_Activity_description == "None") and (Pre_Row_Start_Time == "None") and (Pre_Row_End_Time == "None"):
            if (Row_Personnel_number == "None") and (Row_Date == "None") and (Row_Network_Description == "None") and (Row_Activity == "None") and (Row_Activity_description == "None") and (Row_Start_Time == "None") and (Row_End_Time == "None"):
                Excel_row_No = Row_Index
            else:
                Excel_row_No += 2   # add 2 because of that program works with indexes (start from 0) and find first occurrence of values
                break

        Pre_Row_Personnel_number = Row_Personnel_number
        Pre_Row_Date = Row_Date
        Pre_Row_Network_Description = Row_Network_Description
        Pre_Row_Activity = Row_Activity
        Pre_Row_Activity_description = Row_Activity_description
        Pre_Row_Start_Time = Row_Start_Time
        Pre_Row_End_Time = Row_End_Time

    # Check because of first line
    if Excel_row_No == 0:
        Excel_row_No = 2
    else:
        pass
    
    A_Cell = f"A{Excel_row_No}"
    E_Cell = f"E{Excel_row_No}"
    return A_Cell, E_Cell

# ---------------------------------------------------------- Main Functions ---------------------------------------------------------- #
def Upload(Settings: dict, Configuration: dict, Events: DataFrame, SP_Password: str|None) -> None:
    SP_Team_Current = Settings["0"]["General"]["Downloader"]["Sharepoint"]["Teams"]["My_Team"]
    SP_Link_Current = Settings["0"]["General"]["Downloader"]["Sharepoint"]["Teams"]["Team_Links"][f"{SP_Team_Current}"]

    # Authentication
    s_aut = Authentication.Authentication(Settings=Settings, Configuration=Configuration, SP_Password=SP_Password)

    # Download
    Downloaded = Download_Excel(Settings=Settings, s_aut=s_aut, SP_Link=SP_Link_Current, Type="Current", Name=None)

    if Downloaded == True:
        # Get WorkSheet
        TimeSpent_ws = Get_WorkSheet(Settings=Settings, Sheet_Name="TimeSpent", Type="Current", Name=None)

        # Get Table List from Worksheet
        Table_list = Get_Tables_on_Worksheet(Sheet=TimeSpent_ws)

        # TimeSheets_df
        data_boundary = Table_list[0][1]
        data_boundary = data_boundary.replace("O", "J")
        TimeSheets_df = Get_Table_Data(ws=TimeSpent_ws, data_boundary=data_boundary)
        A_Cell, E_Cell = Time_sheets_Identify_empty_row(TimeSheets_df=TimeSheets_df)

        # TODO --> automatically upload to Sharepoint only to new lines "Paste as text only"
        Elements.Get_MessageBox(Configuration=Configuration, title="Warning Message!", message=f"First Cell: {A_Cell}, {E_Cell} --> Not finished development", icon="warning", fade_in_duration=1, GUI_Level_ID=1)


def Get_Project_and_Activity(Settings: dict, Configuration: dict, SP_Password: str|None) -> None:
    SP_Team_Current = Settings["0"]["General"]["Downloader"]["Sharepoint"]["Teams"]["My_Team"]
    SP_Link_Current = Settings["0"]["General"]["Downloader"]["Sharepoint"]["Teams"]["Team_Links"][f"{SP_Team_Current}"]

    # Authentication
    s_aut = Authentication.Authentication(Settings=Settings, Configuration=Configuration, SP_Password=SP_Password)

    # Download
    Downloaded = Download_Excel(Settings=Settings, s_aut=s_aut, SP_Link=SP_Link_Current, Type="Current", Name=None)
    
    if Downloaded == True:
        Get_Project(Settings=Settings)
        Get_Activity(Settings=Settings)
        
def Get_Project(Settings: dict) -> None:
    SP_File_Name = Settings["0"]["General"]["Downloader"]["Sharepoint"]["File_name"]

    Projects_df = read_excel(io=Data_Functions.Absolute_path(relative_path=f"Operational\\Downloads\\{SP_File_Name}.xlsm"), sheet_name="Projects", usecols="A:C", skiprows=1, nrows=100, header=None)
    Projects_df.drop(columns=[1], inplace=True)
    Projects_df.rename(columns={0: "Project", 2: "Project_Type"}, inplace=True)

    Projects_df = Projects_df.T
    Projects_dict = Projects_df.to_dict()

    # Save to Settings.json
    Data_Functions.Save_Value(Settings=Settings, Configuration=None, Variable=None, File_Name="Settings", JSON_path=["0", "Event_Handler", "Project", "Project_List"], Information=Projects_dict)
    
def Get_Activity(Settings: dict) -> None:
    SP_File_Name = Settings["0"]["General"]["Downloader"]["Sharepoint"]["File_name"]

    Activities_df = read_excel(io=Data_Functions.Absolute_path(relative_path=f"Operational\\Downloads\\{SP_File_Name}.xlsm"), sheet_name="Activity", usecols="A:B", skiprows=1, nrows=100, header=None)
    Column_List = Activities_df[1].to_list()
    Empty_line_index = Column_List.index("Activity")
    Activities_df = Activities_df.iloc[Empty_line_index + 1:]
    Activities_df.reset_index(inplace=True)
    Activities_df.drop(columns=["index"], inplace=True)
    Activities_df.rename(columns={0: "Project_Type", 1: "Activity"}, inplace=True)

    Activity_list = list(set(Activities_df["Activity"]))
    Activity_list.sort()

    Project_Type_list = list(set(Activities_df["Project_Type"]))
    Project_Type_list.sort()

    # Dictionary creation
    Activity_by_Type_dict = {}
    Counter = 0
    for Project_Type in Project_Type_list:
        mask =  Activities_df["Project_Type"] == Project_Type
        Filtered_Df = Activities_df[mask]
        Activity_by_Type_list = Filtered_Df["Activity"].to_list()
        Activity_by_Type_list.sort()

        Activity_by_Type_dict[Counter] = {
            "Project_Type": Project_Type,
            "Activity": Activity_by_Type_list
        }
        Counter += 1

    # Save to Settings.json
    Data_Functions.Save_Value(Settings=Settings, Configuration=None, Variable=None, File_Name="Settings", JSON_path=["0", "Event_Handler", "Activity", "Activity_List"], Information=Activity_list)
    Data_Functions.Save_Value(Settings=Settings, Configuration=None, Variable=None, File_Name="Settings", JSON_path=["0", "Event_Handler", "Activity", "Activity_by_Type_dict"], Information=Activity_by_Type_dict)